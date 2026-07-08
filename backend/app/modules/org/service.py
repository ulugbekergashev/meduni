"""Создание пользователей: общая логика для ручного создания и XLSX-импорта."""
import secrets

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.core.errors import ApiError
from app.core.security import hash_password
from app.modules.auth.models import ROLES, User
from app.modules.org.models import GroupMembership, StudentGroup, TeacherProfile
from app.modules.org.schemas import ImportReport, UserCreate


def create_user(db: Session, data: UserCreate) -> User:
    if data.role not in ROLES:
        raise ApiError(422, "bad_role", "Rol notoʻgʻri", "Неверная роль")
    email = data.email.lower()
    if db.scalar(select(User).where(User.email == email)):
        raise ApiError(409, "email_taken", "Bu email band", "Email уже занят")

    user = User(
        role=data.role,
        full_name=data.full_name,
        email=email,
        phone=data.phone,
        password_hash=hash_password(data.password),
        locale=data.locale if data.locale in ("uz", "ru") else "uz",
    )
    db.add(user)
    db.flush()

    if data.role == "student" and data.group_id is not None:
        db.add(GroupMembership(student_id=user.id, group_id=data.group_id))
    if data.role == "teacher":
        db.add(TeacherProfile(user_id=user.id, department_id=data.department_id))
    return user


def import_users_xlsx(db: Session, content: bytes) -> ImportReport:
    """Колонки XLSX (первая строка — заголовки): full_name, email, role,
    phone?, locale?, password?, group? (имя группы для студентов).
    Пустой password → генерируется случайный (пишется в отчёт)."""
    import io

    wb = load_workbook(io.BytesIO(content), read_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    try:
        header = [str(c).strip().lower() if c else "" for c in next(rows)]
    except StopIteration:
        raise ApiError(422, "empty_file", "Fayl boʻsh", "Файл пуст")

    def col(row: tuple, name: str) -> str | None:
        if name not in header:
            return None
        value = row[header.index(name)]
        return str(value).strip() if value is not None else None

    groups_by_name = {g.name.lower(): g.id for g in db.scalars(select(StudentGroup)).all()}
    created, errors = 0, []

    for i, row in enumerate(rows, start=2):
        full_name, email, role = col(row, "full_name"), col(row, "email"), col(row, "role")
        if not (full_name and email and role):
            if any(v is not None for v in row):
                errors.append(f"строка {i}: нет full_name/email/role")
            continue
        group_id = None
        group_name = col(row, "group")
        if group_name:
            group_id = groups_by_name.get(group_name.lower())
            if group_id is None:
                errors.append(f"строка {i}: группа «{group_name}» не найдена")
                continue
        password = col(row, "password") or secrets.token_urlsafe(8)
        try:
            create_user(db, UserCreate(
                role=role.lower(), full_name=full_name, email=email,
                phone=col(row, "phone"), locale=col(row, "locale") or "uz",
                password=password, group_id=group_id,
            ))
            created += 1
            if not col(row, "password"):
                errors.append(f"строка {i}: {email} — сгенерирован пароль {password}")
        except ApiError as e:
            errors.append(f"строка {i}: {e.detail['message_ru']} ({email})")

    db.commit()
    return ImportReport(created=created, errors=errors)
